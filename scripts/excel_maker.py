#Script_Main_excel_creater0.12 Added code to check and add original vuln class to excel
import json
import os
import sys
from openpyxl import load_workbook
import pandas as pd
import coverage_json
import vulnerabilities_json

filename = sys.argv[1]
Contractname = sys.argv[2] 
fuzzer = sys.argv[4]
ID = sys.argv[5]
root = sys.argv[6]
vulnfi = sys.argv[7]
Vulnerability_detected = "No"
vulnerability = ""
updated_vuln_detected = []


# Error management
if os.path.isfile("coverage_json.json") == False or os.stat("coverage_json.json").st_size == 0 :
    coverage_json.coverage_json_maker()
if os.path.isfile("vuln_json.json") == False or os.stat("vuln_json.json").st_size == 0:
    vulnerabilities_json.vuln_Jsonmaker()


vuln_file = sys.argv[3]
with open(vuln_file) as f:
    lines = f.readlines()
    vulnerability = lines[0]
print(vulnerability)
Vuln_file= "vuln_json.json"

with open(Vuln_file, 'r',encoding="utf-8") as f:
    vuln_json = json.load(f)

for i in vuln_json["Vulnerabilities_detected"]:
    if i == vulnerability:
        Vulnerability_detected = "Yes"


print(Vulnerability_detected)
print(vuln_json["Vulnerabilities_detected"])

if fuzzer == "Confuzzius":
    Dict_vuln = dict({"BlockStateDep" : "Block dependency detected", "Arbitrary": "Arbitrary memory access detected", "Assertion":"Assertion failure detected", "Overflow":"Integer overflow detected", "Underflow":"Integer underflow detected", "Reentrancy":"Reentrancy detected","Transaction":"Transaction order dependency detected","UnhandledException":"Unhandled exception detected","DangerousDelegatecall":"Unsafe delegatecall detected","Leaking":"Leaking ether detected","Locking":"Locking ether detected","Suicidal":"Unprotected selfdestruct detected"})
elif fuzzer =="ILF":
    Dict_vuln = dict({"BlockStateDep": "Block dependency detected","DangerousDelegatecall" :"Unsafe delegatecall detected" , "Leaking": "Leaking ether detected" , "Locking": "Locking ether detected", "Suicidal": "Unprotected selfdestruct detected","UnhandledException": "Unhandled exception detected" , "Reentrancy" : "Reentrancy detected" })
elif fuzzer =="sFuzz":
    Dict_vuln = dict({"gasless":"gasless send detected","Overflow":"Integer overflow/ Underflow detected","BlockStateDep": "timestamp dependency/block number dependency","DangerousDelegatecall" :"Unsafe delegatecall detected"  , "Locking": "freezing ether detected","UnhandledException": "exception disorder detected" , "Reentrancy" : "Reentrancy detected" })
for key, value in Dict_vuln.items():
    if key in vuln_json["Vulnerabilities_detected"] and key != vulnerability:
        updated_vuln_detected.append(value)



#Check vuln class
vulns_class= ["Data","Description","Environment","Interaction","Interface","Logic","Performance","Security","Standard"]
class_vuln = ""
counter = 0
if  fuzzer== "ILF":
    counter = 5
elif fuzzer == "Confuzzius":
    counter = 3
else:
    counter = 6


for vulnclass in vulns_class:
    if vulnclass == vulnfi.split("/")[counter]:
        class_vuln = vulnclass
        
        
        
        
        
        
        
        
File_path= "coverage_json.json"

with open(File_path, 'r',encoding="utf-8") as f:
    json_data = json.load(f)

json_data["Contract_File_name"] = filename
json_data["Contract_Name"] = Contractname
#Vulnerabilities_detected = ",".join(vuln_json["Vulnerabilities_detected"])
Vulnerabilities_detected = ",".join(updated_vuln_detected)
json_data["Vulnerability_detected"] = Vulnerability_detected
json_data["Vulnerabilities_detected"] = Vulnerabilities_detected
json_data["Vulnerability_original"]= vulnerability
json_data["Vuln_class"]=class_vuln

column_names= ["Contract_File_name","Contract_Name","Code_Coverage","Branch_Coverage","Vulnerability_original","Vuln_class","Vulnerability_detected","Vulnerabilities_detected", "No._of_Transactions","Time_Taken","Time_trigger"]
dataframe=pd.DataFrame(columns=column_names)
data = [json_data]
dataframe = dataframe.append(data,ignore_index=True,sort=False)





def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

Filee=root+"dataset/"+fuzzer+"-"+ID+"-output.xlsx"
append_df_to_excel(Filee,dataframe,header=None,index=False)