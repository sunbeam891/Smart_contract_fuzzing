#PLotter_V0.11 _ Added original_file_loc to top
# Usage: python3 Cov_plotter.py <Contract file name> <Contract Name> <Saving folder> <original file location> <Results_folder> <ID>  <vuln file location> <Fuzzer>
import json
import os
import sys
from decimal import Decimal
import sys
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt_1
import numpy as np


contract_file = sys.argv[1]
contract_name = sys.argv[2]
saving_folder = sys.argv[3]
file_loc_orig = sys.argv[4]
results_folder =sys.argv[5]
ID = sys.argv[6]
vuln_file =  sys.argv [7]
Fuzzer = sys.argv [8]




#Creating json file with contract's original information

original_info = {}
original_info["File_name"] = contract_file
original_info["Contract_name"] = contract_name
original_info["Contract_location"] = file_loc_orig
original_info["vuln_file_location"] = vuln_file
    
    
Original_json_loc = saving_folder + "Original_info.json" 

with open(Original_json_loc, 'w', encoding="utf-8") as file:
        x = json.dumps(original_info, indent=4)
        file.write(x + '\n') 



#PLotting graph for coverage vs time
File = saving_folder+"cov_per_time.json"


#Instruction_Coverage
if os.path.isfile(File) == True and not os.stat(File).st_size == 0 and not Fuzzer == "sFuzz":
    with open(File, 'r',encoding="utf-8") as f:
        vuln_json = json.load(f)
    cov_time = pd.DataFrame(columns=('Cov','time'))
    for i,dictionary in enumerate(vuln_json):
        coverage = dictionary["Code_Coverage"].split(" ")[0]
        if "sec" in dictionary["Time_Taken"]:
            time = float(dictionary["Time_Taken"].split(" ")[0])
        else:
            time = float(dictionary["Time_Taken"])
        cov_time.loc[len(cov_time)]=[float(coverage.replace("%","")),time]    
    plt.plot(cov_time["time"],cov_time["Cov"],marker="o")
    #for i,j in zip(cov_time["time"],cov_time["Cov"]):
        #plt.annotate(str(j), xy=(i, j), xytext=(0,5),textcoords='offset points')
    plt.savefig(saving_folder+"plot_instruction.png")
    plt.clf()

    
    
    
#Branch_Coverage    
if os.path.isfile(File) == True and not os.stat(File).st_size == 0 and not Fuzzer=="ILF":
    with open(File, 'r',encoding="utf-8") as f:
        vuln_json = json.load(f)
    cov_time = pd.DataFrame(columns=('Cov','time'))
    for z,dictionary_br in enumerate(vuln_json):
        coverage = dictionary_br["Branch_Coverage"].split(" ")[0]
        if "sec" in dictionary_br["Time_Taken"]:
            time = float(dictionary_br["Time_Taken"].split(" ")[0])
        else:
            time = float(dictionary_br["Time_Taken"])
        cov_time.loc[len(cov_time)]=[float(coverage.replace("%","")),time]    
    plt.plot(cov_time["time"],cov_time["Cov"],marker="o")
    #for i,j in zip(cov_time["time"],cov_time["Cov"]):
        #plt.annotate(str(j), xy=(i, j), xytext=(0,5),textcoords='offset points')
    plt.savefig(saving_folder+"plot_Branch.png")    
    
    
#Storing time for trigger in excel sheet

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()






File = saving_folder+"trigger_time.json"

if os.path.isfile(File) == True and not os.stat(File).st_size == 0: 
    with open(File, 'r',encoding="utf-8") as f:
        vuln_json = json.load(f)
    Columns_trigger = ["Contract_File_name","Contract_name","vulnerability","time_to_trigger","location"]    
    trigger_df = pd.DataFrame(columns = Columns_trigger) 
    previous_vulns = []
    Vuln_dict = {}
    for i , dictionary in enumerate(vuln_json):
        for i in dictionary["vulnerabilities"]:
            if not i in previous_vulns and not i == "_":
                Vuln_dict[i]=dictionary["Time_Taken"]
                previous_vulns.append(i)
    if not len(Vuln_dict) == 0:
        for vuln,time in Vuln_dict.items():
            df_dict = {}
            df_dict["Contract_File_name"] = contract_file
            df_dict["Contract_name"] =contract_name
            df_dict["vulnerability"] = vuln
            df_dict["time_to_trigger"] = time
            df_dict["location"] = file_loc_orig
            data = [df_dict]
            trigger_df = trigger_df.append(data,ignore_index=True,sort=False)    
    else:
        df_dict = {}
        df_dict["Contract_File_name"] = contract_file
        df_dict["Contract_name"] =contract_name
        df_dict["vulnerability"] = "_"
        df_dict["time_to_trigger"] = "-"
        df_dict["location"] = file_loc_orig
        data = [df_dict]
        trigger_df = trigger_df.append(data,ignore_index=True,sort=False)
    Filee=results_folder+"/time_to_trigger_"+ID+".xlsx"
    append_df_to_excel(Filee,trigger_df,header=None,index=False)    

elif os.path.isfile(File) == False or os.stat(File).st_size == 0: 
    Columns_trigger = ["Contract_File_name","Contract_name","vulnerability","time_to_trigger","location"]    
    trigger_df = pd.DataFrame(columns = Columns_trigger) 
    previous_vulns = []
    Vuln_dict = {}
    df_dict = {}
    df_dict["Contract_File_name"] = contract_file
    df_dict["Contract_name"] =contract_name
    df_dict["vulnerability"] = "_"
    df_dict["time_to_trigger"] = "-"
    df_dict["location"] = file_loc_orig
    data = [df_dict]
    trigger_df = trigger_df.append(data,ignore_index=True,sort=False)
    Filee=results_folder+"/time_to_trigger_"+ID+".xlsx"
    append_df_to_excel(Filee,trigger_df,header=None,index=False)

    
    
    
    
   
    
    
    
    
    
    